"""
报表导出路由 — JSON查询 + Excel文件导出
"""
from urllib.parse import quote
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date, datetime
import io
from database import get_db
from models import (Material, InboundRecord, Requisition, RequisitionItem,
                    AuditLog, User, Category)
from auth import get_current_user, require_admin
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from excel_utils import create_workbook, write_row, apply_fill, add_summary_row, finalize
from openpyxl import Workbook

router = APIRouter(prefix="/api/reports", tags=["报表导出"])


@router.get("/stock")
def report_stock(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """库存报表：按类别汇总"""
    categories = db.query(Category).filter(Category.parent_id == None).all()
    result = {"total_value": 0, "total_items": 0, "categories": []}
    for cat in categories:
        materials = db.query(Material).filter(Material.category_id == cat.id, Material.status == 1).all()
        if not materials:
            continue
        cat_value = sum(float(m.stock_qty * m.unit_price) for m in materials)
        cat_count = len(materials)
        result["categories"].append({"name": cat.name, "count": cat_count, "value": cat_value})
        result["total_value"] += cat_value
        result["total_items"] += cat_count
    return result


@router.get("/consumption")
def report_consumption(months: int = Query(6, ge=1, le=24),
                       user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """消耗趋势：最近N个月，统计已出库/已完成的领用量"""
    from calendar import monthrange
    result = []
    now = datetime.utcnow()
    for i in range(months - 1, -1, -1):
        m = now.month - i
        y = now.year
        while m <= 0:
            m += 12
            y -= 1
        start = datetime(y, m, 1)
        end = datetime(y, m, monthrange(y, m)[1], 23, 59, 59)
        delivered = db.query(Requisition).filter(
            Requisition.delivered_at >= start,
            Requisition.delivered_at <= end,
            Requisition.status.in_(["delivered", "signed", "returned_stock"])
        ).all()
        result.append({
            "month": f"{y}年{m}月",
            "order_count": len(delivered),
            "total_amount": round(sum(float(r.total_amount) for r in delivered), 2)
        })
    return result


@router.get("/inbound-summary")
def report_inbound_summary(start_date: date = Query(None), end_date: date = Query(None),
                           user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """入库汇总报表"""
    q = db.query(InboundRecord).filter(InboundRecord.deleted_at == None)
    if start_date:
        q = q.filter(InboundRecord.inbound_date >= start_date)
    if end_date:
        q = q.filter(InboundRecord.inbound_date <= end_date)
    records = q.all()
    return {
        "total_count": len(records),
        "total_amount": sum(float(r.total_amount) for r in records),
        "total_quantity": sum(r.quantity for r in records),
        "records": [{
            "inbound_code": r.inbound_code, "material_name": r.material.name if r.material else "",
            "batch_no": r.batch_no, "quantity": r.quantity,
            "total_amount": float(r.total_amount), "inbound_date": r.inbound_date,
        } for r in records]
    }


@router.get("/supplier-stats")
def report_supplier_stats(start_date: date = Query(None), end_date: date = Query(None),
                          user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """供应商供货统计"""
    from models import Supplier
    suppliers = db.query(Supplier).filter(Supplier.deleted_at == None).all()
    result = []
    for s in suppliers:
        q = db.query(InboundRecord).filter(InboundRecord.supplier_id == s.id,
                                            InboundRecord.deleted_at == None)
        if start_date: q = q.filter(InboundRecord.inbound_date >= start_date)
        if end_date: q = q.filter(InboundRecord.inbound_date <= end_date)
        records = q.all()
        result.append({
            "supplier_name": s.name, "order_count": len(records),
            "total_amount": sum(float(r.total_amount) for r in records),
        })
    return result


# ═══════════════════════════════════════════════════════
#  Excel 导出：库存报表
# ═══════════════════════════════════════════════════════

@router.get("/export/stock")
def export_stock_excel(
    ids: str = Query(None, description="逗号分隔的耗材ID"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """导出库存报表为 Excel 文件"""
    q = db.query(Material).filter(Material.status == 1)
    if ids:
        id_list = [int(x.strip()) for x in ids.split(",") if x.strip().isdigit()]
        if id_list:
            q = q.filter(Material.id.in_(id_list))
    materials = q.order_by(Material.category_id, Material.name).all()

    headers = ["序号", "耗材编号", "名称", "规格型号", "类别", "单位",
               "单价(元)", "库存数量", "库存价值(元)", "安全下限", "安全上限", "预警状态", "存放位置"]
    widths = [6, 18, 18, 20, 14, 8, 12, 10, 14, 10, 10, 10, 24]
    wb, ws = create_workbook("库存报表", headers, widths)

    total_value = 0.0
    alert_count = {"low": 0, "high": 0, "normal": 0}
    for idx, m in enumerate(materials, 1):
        row = idx + 2
        alert = "正常"
        if m.safety_stock_min and m.stock_qty < m.safety_stock_min:
            alert = "低库存预警"
            alert_count["low"] += 1
        elif m.safety_stock_max and m.stock_qty > m.safety_stock_max:
            alert = "高库存预警"
            alert_count["high"] += 1
        else:
            alert_count["normal"] += 1

        stock_value = float(m.stock_qty * m.unit_price)
        total_value += stock_value

        values = [idx, m.material_code, m.name, m.spec,
                  m.category.name if m.category else "", m.unit,
                  round(float(m.unit_price), 4), m.stock_qty,
                  round(stock_value, 2),
                  m.safety_stock_min or "", m.safety_stock_max or "",
                  alert, m.location]
        write_row(ws, row, values)

        if alert == "低库存预警":
            apply_fill(ws, row, len(headers), "FFD7D7")
        elif alert == "高库存预警":
            apply_fill(ws, row, len(headers), "FFF2CC")

    summary_row = len(materials) + 3
    if materials:
        add_summary_row(ws, summary_row, len(headers),
            f"共 {len(materials)} 种耗材（低预警{alert_count['low']}项 / 高预警{alert_count['high']}项）",
            {8: f"总库存价值: {round(total_value,2)}"})

    filename = f"库存报表_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        finalize(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


# ═══════════════════════════════════════════════════════
#  Excel 导出：供应商供货统计
# ═══════════════════════════════════════════════════════

@router.get("/export/supplier-stats")
def export_supplier_excel(
    start_date: date = Query(None),
    end_date: date = Query(None),
    ids: str = Query(None, description="逗号分隔的供应商ID"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """导出供应商供货统计为 Excel 文件"""
    from models import Supplier
    q = db.query(Supplier).filter(Supplier.deleted_at == None)
    if ids:
        id_list = [int(x.strip()) for x in ids.split(",") if x.strip().isdigit()]
        if id_list:
            q = q.filter(Supplier.id.in_(id_list))
    suppliers = q.order_by(Supplier.name).all()

    headers = ["序号", "供应商编号", "供应商名称", "联系人", "联系电话",
               "合作状态", "供货次数", "供货总金额(元)"]
    widths = [6, 16, 30, 12, 18, 10, 12, 18]
    wb, ws = create_workbook("供应商供货统计", headers, widths)

    total_orders = 0
    total_amt = 0.0
    for idx, s in enumerate(suppliers, 1):
        row = idx + 2
        q = db.query(InboundRecord).filter(
            InboundRecord.supplier_id == s.id,
            InboundRecord.deleted_at == None
        )
        if start_date: q = q.filter(InboundRecord.inbound_date >= start_date)
        if end_date: q = q.filter(InboundRecord.inbound_date <= end_date)
        records = q.all()
        order_count = len(records)
        order_amount = sum(float(r.total_amount) for r in records)

        values = [idx, s.supplier_code, s.name, s.contact_person, s.contact_phone,
                  "合作中" if s.coop_status.value == "active" else "已停用",
                  order_count, round(order_amount, 2)]
        write_row(ws, row, values)
        if s.coop_status.value != "active":
            apply_fill(ws, row, len(headers), "F2F2F2")
        total_orders += order_count
        total_amt += order_amount

    summary_row = len(suppliers) + 3
    if suppliers:
        add_summary_row(ws, summary_row, len(headers),
            f"共 {len(suppliers)} 家供应商",
            {7: f"合计: {total_orders}次", 8: f"合计: {round(total_amt,2)}"})

    filename = f"供应商统计_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        finalize(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


# ═══════════════════════════════════════════════════════
#  Excel 导出：盘点差异
# ═══════════════════════════════════════════════════════

@router.get("/export/discrepancy")
def export_discrepancy_excel(
    ids: str = Query(None, description="逗号分隔的耗材ID"),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    导出盘点差异报表为 Excel 文件。
    盘点差异 = 当前库存 - 理论库存（入库 - 出库 + 归还）
    供实际盘点时核对使用。
    """
    q = db.query(Material).filter(Material.status == 1)
    if ids:
        id_list = [int(x.strip()) for x in ids.split(",") if x.strip().isdigit()]
        if id_list:
            q = q.filter(Material.id.in_(id_list))
    materials = q.order_by(Material.category_id, Material.name).all()

    headers = ["序号", "耗材编号", "名称", "规格型号", "类别", "单位",
               "当前库存", "入库总量", "出库总量", "归还总量",
               "理论库存", "差异", "单价(元)", "差异金额(元)", "存放位置"]
    widths = [6, 18, 18, 16, 12, 8, 10, 10, 10, 10, 10, 8, 12, 14, 20]
    wb, ws = create_workbook("盘点差异报表", headers, widths)

    total_diff_value = 0.0
    for idx, m in enumerate(materials, 1):
        row = idx + 2

        # 入库总量
        inbound_total = db.query(func.coalesce(func.sum(InboundRecord.quantity), 0))\
            .filter(InboundRecord.material_id == m.id, InboundRecord.deleted_at == None).scalar()

        # 出库总量（已签收的）
        delivered = db.query(func.coalesce(func.sum(RequisitionItem.actual_quantity), 0))\
            .join(Requisition)\
            .filter(RequisitionItem.material_id == m.id,
                    Requisition.status.in_(["delivered", "signed", "returned_stock"])).scalar()

        # 归还总量
        returned = db.query(func.coalesce(func.sum(RequisitionItem.return_quantity), 0))\
            .filter(RequisitionItem.material_id == m.id).scalar()

        theoretical = inbound_total - delivered + (returned or 0)
        discrepancy = m.stock_qty - theoretical
        diff_value = discrepancy * float(m.unit_price)
        total_diff_value += abs(diff_value)

        values = [idx, m.material_code, m.name, m.spec,
                  m.category.name if m.category else "", m.unit,
                  m.stock_qty, inbound_total, delivered, returned or 0,
                  theoretical, discrepancy,
                  round(float(m.unit_price), 4),
                  round(diff_value, 2), m.location]
        write_row(ws, row, values)

        if abs(discrepancy) > 0:
            apply_fill(ws, row, len(headers), "FFD7D7")

    summary_row = len(materials) + 3
    if materials:
        add_summary_row(ws, summary_row, len(headers),
            f"盘点差异绝对值合计: {round(total_diff_value,2)} 元",
            {})

    filename = f"盘点差异_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        finalize(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


# ═══════════════════════════════════════════════════════
#  Excel 导出：数据备份（全量导出）
# ═══════════════════════════════════════════════════════

@router.get("/export/backup")
def export_backup_excel(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    数据备份：将所有核心数据导出为 Excel 文件（多Sheet）
    Sheet1=耗材, Sheet2=入库记录, Sheet3=领用申请, Sheet4=归还记录, Sheet5=供应商
    """
    from models import Supplier, ReturnRecord
    wb = Workbook()
    # 删除默认sheet
    wb.remove(wb.active)

    # ── Sheet 1: 耗材 ──
    headers_m = ["序号", "编号", "名称", "规格", "类别", "单位", "单价", "库存", "安全下限", "安全上限", "位置"]
    ws_m = wb.create_sheet("耗材清单")
    for i, h in enumerate(headers_m, 1):
        c = ws_m.cell(row=1, column=i, value=h)
        c.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        c.alignment = Alignment(horizontal="center")
        c.border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    materials = db.query(Material).filter(Material.status == 1).order_by(Material.id).all()
    for idx, m in enumerate(materials, 1):
        write_row(ws_m, idx + 1, [idx, m.material_code, m.name, m.spec,
            m.category.name if m.category else "", m.unit,
            float(m.unit_price), m.stock_qty, m.safety_stock_min or "",
            m.safety_stock_max or "", m.location])

    # ── Sheet 2: 入库记录 ──
    headers_i = ["序号", "入库单号", "耗材", "批次号", "数量", "单价", "金额", "供应商", "日期", "经手人"]
    ws_i = wb.create_sheet("入库记录")
    for i, h in enumerate(headers_i, 1):
        c = ws_i.cell(row=1, column=i, value=h)
        c.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        c.alignment = Alignment(horizontal="center")
        c.border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    inbounds = db.query(InboundRecord).filter(InboundRecord.deleted_at == None).order_by(InboundRecord.id).all()
    for idx, r in enumerate(inbounds, 1):
        write_row(ws_i, idx + 1, [idx, r.inbound_code,
            r.material.name if r.material else "", r.batch_no, r.quantity,
            float(r.unit_price), float(r.total_amount),
            r.supplier.name if r.supplier else "",
            r.inbound_date.strftime("%Y-%m-%d") if r.inbound_date else "",
            r.operator.real_name if r.operator else ""])

    # ── Sheet 3: 领用申请 ──
    headers_r = ["序号", "申请单号", "申请人", "用途", "总金额", "状态", "审批人", "出库时间", "申请时间"]
    ws_r = wb.create_sheet("领用申请")
    for i, h in enumerate(headers_r, 1):
        c = ws_r.cell(row=1, column=i, value=h)
        c.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        c.alignment = Alignment(horizontal="center")
        c.border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    reqs = db.query(Requisition).order_by(Requisition.id).all()
    for idx, r in enumerate(reqs, 1):
        write_row(ws_r, idx + 1, [idx, r.req_code,
            r.applicant.real_name if r.applicant else "", r.purpose,
            float(r.total_amount), r.status.value,
            r.approver.real_name if r.approver else "",
            r.delivered_at.strftime("%Y-%m-%d") if r.delivered_at else "",
            r.created_at.strftime("%Y-%m-%d") if r.created_at else ""])

    # ── Sheet 4: 归还记录 ──
    headers_ret = ["序号", "归还单号", "领用单号", "耗材", "数量", "状态", "逾期天数", "罚款", "日期"]
    ws_ret = wb.create_sheet("归还记录")
    for i, h in enumerate(headers_ret, 1):
        c = ws_ret.cell(row=1, column=i, value=h)
        c.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        c.alignment = Alignment(horizontal="center")
        c.border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    returns = db.query(ReturnRecord).order_by(ReturnRecord.id).all()
    for idx, r in enumerate(returns, 1):
        mat = db.query(Material).filter(Material.id == r.material_id).first()
        write_row(ws_ret, idx + 1, [idx, r.return_code,
            r.requisition.req_code if r.requisition else "",
            mat.name if mat else "", r.return_quantity,
            r.return_condition.value, r.overdue_days, float(r.fine_amount),
            r.return_date.strftime("%Y-%m-%d") if r.return_date else ""])

    # ── Sheet 5: 供应商 ──
    headers_s = ["序号", "编号", "名称", "联系人", "电话", "合作状态"]
    ws_s = wb.create_sheet("供应商")
    for i, h in enumerate(headers_s, 1):
        c = ws_s.cell(row=1, column=i, value=h)
        c.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        c.alignment = Alignment(horizontal="center")
        c.border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    suppliers = db.query(Supplier).filter(Supplier.deleted_at == None).order_by(Supplier.id).all()
    for idx, s in enumerate(suppliers, 1):
        write_row(ws_s, idx + 1, [idx, s.supplier_code, s.name, s.contact_person, s.contact_phone, s.coop_status.value])

    # 列宽自动调整
    for ws in [ws_m, ws_i, ws_r, ws_ret, ws_s]:
        ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"数据备份_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )
