"""
物品归还管理路由：归还登记 / 逾期预览 / Excel导出 / 罚款管理
"""
import io
from datetime import datetime, date
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import and_, func

from database import get_db
from models import (
    ReturnRecord, ReturnCondition, Requisition, RequisitionItem,
    Material, AuditLog, User, ReqStatus
)
from schemas import (
    ReturnCreate, ReturnRecordResponse, OverduePreviewRequest,
    OverduePreviewResponse, FineUpdate, PaginatedResponse, MessageResponse
)
from auth import get_current_user, require_admin
from config import (
    RETURN_FINE_RATE, RETURN_OVERDUE_GRACE_DAYS,
    RETURN_DAMAGE_RATE, RETURN_LOSS_RATE, RETURN_MAX_BORROW_DAYS
)

router = APIRouter(prefix="/api/returns", tags=["物品归还"])


# ═══════════════════════════════════════════════════════
#  工具函数
# ═══════════════════════════════════════════════════════

def generate_return_code(db: Session) -> str:
    """生成归还编号: GH + YYYYMMDD + 4位流水"""
    today = datetime.utcnow().strftime("%Y%m%d")
    last = db.query(ReturnRecord).filter(
        ReturnRecord.return_code.like(f"GH{today}%")
    ).order_by(ReturnRecord.return_code.desc()).first()
    seq = int(last.return_code[-4:]) + 1 if last else 1
    return f"GH{today}{seq:04d}"


def calc_fine(
    delivered_at: datetime,
    return_condition: str,
    unit_price: float,
    return_quantity: int
) -> tuple[int, float, float]:
    """
    计算逾期罚款
    返回: (overdue_days, fine_rate, fine_amount)
    """
    overdue_days = 0
    fine_amount = 0.0

    if delivered_at:
        delta = (datetime.utcnow() - delivered_at).days
        overdue_days = max(0, delta - RETURN_MAX_BORROW_DAYS - RETURN_OVERDUE_GRACE_DAYS)

    if return_condition == "lost":
        # 丢失：按原价全额赔偿
        fine_amount = float(unit_price) * return_quantity * RETURN_LOSS_RATE
        fine_rate = float(unit_price) * RETURN_LOSS_RATE
    elif return_condition == "damaged":
        # 损坏：按原价30%赔偿 + 逾期费
        fine_amount = float(unit_price) * return_quantity * RETURN_DAMAGE_RATE
        fine_rate = float(unit_price) * RETURN_DAMAGE_RATE
    else:
        fine_rate = RETURN_FINE_RATE

    # 完好但逾期：加收逾期费
    if return_condition == "good" and overdue_days > 0:
        fine_amount = overdue_days * RETURN_FINE_RATE * return_quantity
    elif return_condition == "damaged" and overdue_days > 0:
        fine_amount += overdue_days * RETURN_FINE_RATE * return_quantity

    return overdue_days, fine_rate, round(fine_amount, 2)


# ═══════════════════════════════════════════════════════
#  1. 逾期预览（不产生实际记录）
# ═══════════════════════════════════════════════════════

@router.post("/overdue-preview", response_model=OverduePreviewResponse)
def preview_overdue(
    data: OverduePreviewRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    预览逾期罚款：传入领用申请ID，返回每个明细项的预计逾期天数和罚款金额。
    不会产生数据库记录，仅用于归还前的参考信息。
    """
    req = db.query(Requisition).options(
        joinedload(Requisition.items).joinedload(RequisitionItem.material)
    ).filter(Requisition.id == data.requisition_id).first()

    if not req:
        raise HTTPException(status_code=404, detail="领用申请不存在")
    if req.status not in [ReqStatus.delivered, ReqStatus.signed]:
        raise HTTPException(status_code=400, detail="只有已出库或已签收的申请可以预览归还")

    delivered_at = req.delivered_at
    preview_items = []
    total_fine = 0.0

    for item in req.items:
        mat = item.material
        remaining = (item.actual_quantity or item.req_quantity) - (item.return_quantity or 0)
        if remaining <= 0:
            continue

        overdue_days, fine_rate, fine_amount = calc_fine(
            delivered_at, "good", float(mat.unit_price), 1
        )

        preview_items.append({
            "item_id": item.id,
            "item_code": item.item_code,
            "material_name": mat.name if mat else "",
            "material_spec": mat.spec if mat else "",
            "unit_price": float(mat.unit_price) if mat else 0,
            "remaining_quantity": remaining,
            "overdue_days": overdue_days,
            "fine_rate_per_day": fine_rate,
            "estimated_fine_per_unit": round(fine_amount, 2),
            "estimated_fine_total": round(fine_amount * remaining, 2),
        })
        total_fine += fine_amount * remaining

    return OverduePreviewResponse(
        requisition_id=req.id,
        req_code=req.req_code,
        applicant_name=req.applicant.real_name if req.applicant else "",
        delivered_at=delivered_at,
        overdue_days=max(
            (datetime.utcnow() - delivered_at).days - RETURN_MAX_BORROW_DAYS - RETURN_OVERDUE_GRACE_DAYS
            if delivered_at else 0, 0
        ),
        items=preview_items,
        total_fine=round(total_fine, 2),
    )


# ═══════════════════════════════════════════════════════
#  2. 归还列表（分页 + 多条件筛选）
# ═══════════════════════════════════════════════════════

@router.get("", response_model=PaginatedResponse)
def list_returns(
    keyword: str = Query(None, description="搜索关键词(归还编号/耗材名/申请人)"),
    return_condition: str = Query(None, description="归还状态: good/damaged/lost"),
    requisition_id: int = Query(None, description="领用申请ID"),
    start_date: date = Query(None, description="归还开始日期"),
    end_date: date = Query(None, description="归还截止日期"),
    fine_paid: int = Query(None, description="罚款状态: 0=未缴, 1=已缴"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    归还记录列表，支持按关键词、归还状态、日期范围、罚款状态筛选。
    教师只能看到自己申请的归还记录。
    """
    q = db.query(ReturnRecord)

    # 教师只能看自己的归还记录
    if user.role.value == "teacher":
        q = q.join(Requisition, ReturnRecord.requisition_id == Requisition.id)
        q = q.filter(Requisition.applicant_id == user.id)

    if keyword:
        q = q.filter(
            ReturnRecord.return_code.contains(keyword) |
            ReturnRecord.remark.contains(keyword)
        )
    if return_condition:
        q = q.filter(ReturnRecord.return_condition == return_condition)
    if requisition_id is not None:
        q = q.filter(ReturnRecord.requisition_id == requisition_id)
    if start_date:
        q = q.filter(ReturnRecord.return_date >= start_date)
    if end_date:
        q = q.filter(ReturnRecord.return_date <= end_date)
    if fine_paid is not None:
        q = q.filter(ReturnRecord.fine_paid == fine_paid)

    total = q.count()
    records = q.order_by(ReturnRecord.created_at.desc())\
               .offset((page - 1) * page_size).limit(page_size).all()

    items = []
    for r in records:
        mat = db.query(Material).filter(Material.id == r.material_id).first()
        rit = db.query(RequisitionItem).filter(RequisitionItem.id == r.requisition_item_id).first()
        items.append({
            "id": r.id,
            "return_code": r.return_code,
            "requisition_id": r.requisition_id,
            "req_code": r.requisition.req_code if r.requisition else None,
            "item_code": rit.item_code if rit else None,
            "material_name": mat.name if mat else "",
            "material_spec": mat.spec if mat else "",
            "return_quantity": r.return_quantity,
            "return_condition": r.return_condition.value if r.return_condition else "",
            "overdue_days": r.overdue_days,
            "fine_rate": float(r.fine_rate),
            "fine_amount": float(r.fine_amount),
            "fine_paid": r.fine_paid,
            "handler_name": r.handler.real_name if r.handler else "",
            "return_date": r.return_date,
            "remark": r.remark,
            "created_at": r.created_at,
        })

    return PaginatedResponse(
        items=items, total=total, page=page, page_size=page_size,
        total_pages=(total + page_size - 1) // page_size
    )


# ═══════════════════════════════════════════════════════
#  3. 归还详情
# ═══════════════════════════════════════════════════════

@router.get("/{return_id}")
def get_return_detail(
    return_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取归还记录详情"""
    r = db.query(ReturnRecord).filter(ReturnRecord.id == return_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="归还记录不存在")

    mat = db.query(Material).filter(Material.id == r.material_id).first()
    rit = db.query(RequisitionItem).filter(RequisitionItem.id == r.requisition_item_id).first()

    return {
        "id": r.id,
        "return_code": r.return_code,
        "requisition_id": r.requisition_id,
        "req_code": r.requisition.req_code if r.requisition else None,
        "requisition_item_id": r.requisition_item_id,
        "item_code": rit.item_code if rit else None,
        "material_id": r.material_id,
        "material_name": mat.name if mat else "",
        "material_spec": mat.spec if mat else "",
        "material_unit": mat.unit if mat else "",
        "return_quantity": r.return_quantity,
        "return_condition": r.return_condition.value if r.return_condition else "",
        "overdue_days": r.overdue_days,
        "fine_rate": float(r.fine_rate),
        "fine_amount": float(r.fine_amount),
        "fine_paid": r.fine_paid,
        "handler_name": r.handler.real_name if r.handler else "",
        "return_date": r.return_date.isoformat() if r.return_date else None,
        "remark": r.remark,
        "created_at": r.created_at.isoformat() if r.created_at else None,
    }


# ═══════════════════════════════════════════════════════
#  4. 核心：创建归还记录（事务安全）
# ═══════════════════════════════════════════════════════

@router.post("", response_model=MessageResponse)
def create_return(
    data: ReturnCreate,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    执行归还登记。

    业务流程：
    1. 校验领用申请状态（必须 delivered/signed）
    2. 逐项校验归还数量 ≤ 剩余可归还量
    3. 计算逾期天数和罚款金额
    4. 更新 requisition_items.return_quantity
    5. 完好的耗材恢复库存：material.stock_qty += return_quantity
    6. 所有项归还完毕 → requisition.status = returned_stock
    7. 记录审计日志
    8. 全部在一个事务内完成，任一步骤失败自动回滚。
    """
    # ── 1. 校验领用申请 ──
    req = db.query(Requisition).options(
        joinedload(Requisition.items).joinedload(RequisitionItem.material)
    ).filter(Requisition.id == data.requisition_id).first()

    if not req:
        raise HTTPException(status_code=404, detail="领用申请不存在")
    if req.status not in [ReqStatus.delivered, ReqStatus.signed]:
        raise HTTPException(status_code=400,
            detail=f"当前申请状态为 {req.status.value}，只有已出库或已签收的申请可以归还")

    # ── 2. 逐项校验与处理 ──
    item_map = {it.id: it for it in req.items}
    return_codes = []

    for entry in data.items:
        item = item_map.get(entry.requisition_item_id)
        if not item:
            raise HTTPException(status_code=400,
                detail=f"明细ID {entry.requisition_item_id} 不属于此申请")

        # 可归还量 = 实出数量 - 已归还量
        actual = item.actual_quantity or item.req_quantity
        already_returned = item.return_quantity or 0
        remaining = actual - already_returned

        if entry.return_quantity > remaining:
            raise HTTPException(status_code=400,
                detail=f"耗材 '{item.material.name}' 最多可归还 {remaining}，当前提交 {entry.return_quantity}")

        if entry.return_quantity <= 0:
            raise HTTPException(status_code=400,
                detail=f"归还数量必须大于0（明细ID: {entry.requisition_item_id}）")

        # ── 3. 计算逾期罚款 ──
        mat = item.material
        overdue_days, fine_rate, fine_amount = calc_fine(
            req.delivered_at, entry.return_condition,
            float(mat.unit_price), entry.return_quantity
        )

        # ── 4. 生成归还编号并创建记录 ──
        code = generate_return_code(db)
        return_codes.append(code)

        rr = ReturnRecord(
            return_code=code,
            requisition_id=req.id,
            requisition_item_id=item.id,
            material_id=item.material_id,
            return_quantity=entry.return_quantity,
            return_date=datetime.utcnow(),
            return_condition=ReturnCondition(entry.return_condition),
            overdue_days=overdue_days,
            fine_rate=fine_rate,
            fine_amount=fine_amount,
            fine_paid=0,
            handler_id=user.id,
            remark=data.remark,
        )
        db.add(rr)

        # ── 5. 更新明细归还量 ──
        item.return_quantity = already_returned + entry.return_quantity
        item.returned_at = datetime.utcnow()

        # ── 6. 完好的归还库存 ──
        if entry.return_condition == "good":
            mat.stock_qty += entry.return_quantity

        # ── 审计日志 ──
        cond_label = {"good": "完好", "damaged": "损坏", "lost": "丢失"}
        log = AuditLog(
            user_id=user.id, username=user.username,
            action_type="return",
            action_detail=(
                f"归还耗材：{mat.name} {mat.spec} ×{entry.return_quantity}{mat.unit}"
                f"（{cond_label.get(entry.return_condition, entry.return_condition)}，"
                f"逾期{overdue_days}天，罚金{fine_amount}元）"
                f"（归还单号：{code}）"
            ),
            target_type="return", target_id=rr.id, ip_address="127.0.0.1"
        )
        db.add(log)

    # ── 7. 检查是否所有项都已归还完毕 ──
    all_returned = True
    for it in req.items:
        actual = it.actual_quantity or it.req_quantity
        if (it.return_quantity or 0) < actual:
            all_returned = False
            break

    if all_returned:
        old_status = req.status
        req.status = ReqStatus.returned_stock
        log = AuditLog(
            user_id=user.id, username=user.username,
            action_type="return_complete",
            action_detail=f"全部归还完毕，申请单 {req.req_code} 状态 {old_status.value} → returned_stock",
            target_type="requisition", target_id=req.id, ip_address="127.0.0.1"
        )
        db.add(log)

    # ── 8. 提交事务 ──
    db.commit()

    codes_str = "、".join(return_codes)
    return MessageResponse(
        message=f"归还登记成功，归还单号：{codes_str}",
        success=True
    )


# ═══════════════════════════════════════════════════════
#  5. 更新罚款缴纳状态
# ═══════════════════════════════════════════════════════

@router.put("/{return_id}/fine", response_model=MessageResponse)
def update_fine_status(
    return_id: int,
    data: FineUpdate,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """标记罚款已缴纳"""
    r = db.query(ReturnRecord).filter(ReturnRecord.id == return_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="归还记录不存在")

    r.fine_paid = data.fine_paid
    log = AuditLog(
        user_id=user.id, username=user.username,
        action_type="fine_update",
        action_detail=f"{'标记已缴' if data.fine_paid else '标记未缴'}罚款 {r.fine_amount}元（归还单号：{r.return_code}）",
        target_type="return", target_id=r.id, ip_address="127.0.0.1"
    )
    db.add(log)
    db.commit()

    return MessageResponse(
        message=f"罚款状态已更新为：{'已缴纳' if data.fine_paid else '未缴纳'}",
        success=True
    )


# ═══════════════════════════════════════════════════════
#  6. 归还统计（工作台用）
# ═══════════════════════════════════════════════════════

@router.get("/stats/summary")
def return_stats(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """归还统计摘要：总归还数、逾期数、未缴罚款总额"""
    total_returns = db.query(ReturnRecord).count()
    total_overdue = db.query(ReturnRecord).filter(ReturnRecord.overdue_days > 0).count()
    total_damaged = db.query(ReturnRecord).filter(ReturnRecord.return_condition == "damaged").count()
    total_lost = db.query(ReturnRecord).filter(ReturnRecord.return_condition == "lost").count()
    unpaid_fine = db.query(func.sum(ReturnRecord.fine_amount))\
                     .filter(ReturnRecord.fine_paid == 0).scalar() or 0

    return {
        "total_returns": total_returns,
        "total_overdue": total_overdue,
        "total_damaged": total_damaged,
        "total_lost": total_lost,
        "unpaid_fine": float(unpaid_fine),
    }


# ═══════════════════════════════════════════════════════
#  7. Excel 导出
# ═══════════════════════════════════════════════════════

@router.get("/export/excel")
def export_returns_excel(
    keyword: str = Query(None),
    return_condition: str = Query(None),
    requisition_id: int = Query(None),
    start_date: date = Query(None),
    end_date: date = Query(None),
    fine_paid: int = Query(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    导出归还明细为 Excel (.xlsx) 文件。
    支持按归还状态、日期范围、罚款状态、关键词等条件筛选。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── 构建查询 ──
    q = db.query(ReturnRecord)

    if user.role.value == "teacher":
        q = q.join(Requisition, ReturnRecord.requisition_id == Requisition.id)
        q = q.filter(Requisition.applicant_id == user.id)

    if keyword:
        q = q.filter(ReturnRecord.return_code.contains(keyword) |
                     ReturnRecord.remark.contains(keyword))
    if return_condition:
        q = q.filter(ReturnRecord.return_condition == return_condition)
    if requisition_id is not None:
        q = q.filter(ReturnRecord.requisition_id == requisition_id)
    if start_date:
        q = q.filter(ReturnRecord.return_date >= start_date)
    if end_date:
        q = q.filter(ReturnRecord.return_date <= end_date)
    if fine_paid is not None:
        q = q.filter(ReturnRecord.fine_paid == fine_paid)

    records = q.order_by(ReturnRecord.created_at.desc()).all()

    # ── 创建 Excel ──
    wb = Workbook()
    ws = wb.active
    ws.title = "归还明细"

    # 样式定义
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="微软雅黑", size=10)
    cell_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    damaged_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    lost_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    overdue_fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
    paid_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # 标题行
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = f"zachary · 物品归还明细报表（导出时间：{datetime.utcnow().strftime('%Y-%m-%d %H:%M')}）"
    title_cell.font = Font(name="微软雅黑", bold=True, size=14, color="2B579A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # 筛选条件行
    filters = []
    if start_date: filters.append(f"开始: {start_date}")
    if end_date: filters.append(f"截止: {end_date}")
    if return_condition: filters.append(f"状态: {return_condition}")
    if fine_paid is not None: filters.append(f"罚款: {'已缴' if fine_paid else '未缴'}")
    ws.merge_cells("A2:N2")
    ws["A2"].value = "筛选条件：" + ("、".join(filters) if filters else "全部")
    ws["A2"].font = Font(name="微软雅黑", size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    # 表头
    headers = [
        "序号", "归还单号", "领用单号", "明细编号", "耗材名称", "规格型号",
        "归还数量", "归还状态", "归还日期", "逾期天数",
        "罚金标准(元)", "罚款金额(元)", "罚款状态", "经手人"
    ]
    col_widths = [6, 20, 20, 18, 22, 20, 10, 10, 22, 10, 14, 14, 10, 12]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 28

    # 数据行
    cond_labels = {"good": "完好", "damaged": "损坏", "lost": "丢失"}
    total_fine = 0.0
    total_unpaid = 0.0

    for idx, r in enumerate(records, 1):
        row = idx + 3
        mat = db.query(Material).filter(Material.id == r.material_id).first()
        rit = db.query(RequisitionItem).filter(RequisitionItem.id == r.requisition_item_id).first()

        values = [
            idx,
            r.return_code,
            r.requisition.req_code if r.requisition else "",
            rit.item_code if rit else "",
            mat.name if mat else "",
            mat.spec if mat else "",
            r.return_quantity,
            cond_labels.get(r.return_condition.value, r.return_condition.value) if r.return_condition else "",
            r.return_date.strftime("%Y-%m-%d %H:%M") if r.return_date else "",
            r.overdue_days,
            round(float(r.fine_rate), 2),
            round(float(r.fine_amount), 2),
            "已缴纳" if r.fine_paid else "未缴纳",
            r.handler.real_name if r.handler else "",
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

        # 条件着色
        cond = r.return_condition.value if r.return_condition else ""
        if cond == "damaged":
            for c in range(1, 15):
                ws.cell(row=row, column=c).fill = damaged_fill
        elif cond == "lost":
            for c in range(1, 15):
                ws.cell(row=row, column=c).fill = lost_fill
        elif r.overdue_days > 0:
            for c in range(1, 15):
                ws.cell(row=row, column=c).fill = overdue_fill

        if r.fine_paid:
            ws.cell(row=row, column=13).fill = paid_fill

        total_fine += float(r.fine_amount)
        if not r.fine_paid:
            total_unpaid += float(r.fine_amount)

    # 汇总行
    summary_row = len(records) + 4
    ws.merge_cells(f"A{summary_row}:F{summary_row}")
    ws.cell(row=summary_row, column=1, value=f"共 {len(records)} 条记录").font = Font(name="微软雅黑", bold=True, size=10)
    ws.cell(row=summary_row, column=1).alignment = cell_align
    ws.cell(row=summary_row, column=1).border = thin_border
    for c in range(2, 7):
        ws.cell(row=summary_row, column=c).border = thin_border

    ws.cell(row=summary_row, column=7, value="合计：").font = Font(name="微软雅黑", bold=True, size=10)
    ws.cell(row=summary_row, column=7).alignment = cell_align
    ws.cell(row=summary_row, column=7).border = thin_border

    for c in range(8, 11):
        ws.cell(row=summary_row, column=c).border = thin_border

    ws.cell(row=summary_row, column=11, value=round(total_fine, 2)).font = Font(name="微软雅黑", bold=True, size=10)
    ws.cell(row=summary_row, column=11).alignment = cell_align
    ws.cell(row=summary_row, column=11).border = thin_border

    ws.cell(row=summary_row, column=12, value=round(total_unpaid, 2)).font = Font(name="微软雅黑", bold=True, size=10, color="CC0000")
    ws.cell(row=summary_row, column=12).alignment = cell_align
    ws.cell(row=summary_row, column=12).border = thin_border

    for c in range(13, 15):
        ws.cell(row=summary_row, column=c).border = thin_border

    # 冻结表头
    ws.freeze_panes = "A4"

    # 输出到流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    filename = f"归还明细_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )
