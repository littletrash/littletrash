"""
Excel 导出公共工具
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io


def create_workbook(title: str, headers: list, col_widths: list = None):
    """创建带标准样式的 Excel 工作簿"""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # sheet name limit

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value=f"zachary · {title}（导出时间：{datetime.utcnow().strftime('%Y-%m-%d %H:%M')}）")
    c.font = Font(name="微软雅黑", bold=True, size=13, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # 表头
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[2].height = 26

    # 列宽
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    return wb, ws


def thin_border():
    return Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )


def cell_style(font_size=10, bold=False, color="000000", align="center"):
    return Font(name="微软雅黑", size=font_size, bold=bold, color=color), \
           Alignment(horizontal=align, vertical="center")


def write_row(ws, row, values, bold=False, color=None, fill_color=None):
    """写入一行数据"""
    font, align = cell_style(bold=bold, color=color or "000000")
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = font
        c.alignment = align
        c.border = thin_border()
        if fill_color:
            c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")


def apply_fill(ws, row, col_count, fill_color):
    """对整行着色"""
    for c in range(1, col_count + 1):
        ws.cell(row=row, column=c).fill = PatternFill(
            start_color=fill_color, end_color=fill_color, fill_type="solid"
        )


def add_summary_row(ws, row, col_count, summary_text, summary_values: dict):
    """添加汇总行（字典 key=列索引, value=汇总值）"""
    # 先取消已有合并，防止 MergedCell 冲突
    try:
        ws.unmerge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    except Exception:
        pass
    # 在 summary_values 之外的列范围合并
    val_cols = set(summary_values.keys())
    first_free = min(val_cols) - 1 if val_cols else col_count
    if first_free > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=first_free)
    ws.cell(row=row, column=1, value=summary_text)
    ws.cell(row=row, column=1).font = Font(name="微软雅黑", bold=True, size=10, color="1F4E79")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    # 给所有格子加边框
    for c in range(1, col_count + 1):
        try:
            ws.cell(row=row, column=c).border = thin_border()
        except AttributeError:
            pass
    # 写入汇总值
    for col_idx, value in summary_values.items():
        ws.cell(row=row, column=col_idx, value=value)
        ws.cell(row=row, column=col_idx).font = Font(name="微软雅黑", bold=True, size=10, color="CC0000")
        ws.cell(row=row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")


def finalize(wb) -> io.BytesIO:
    """输出为流"""
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
